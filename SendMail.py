import smtplib # Lib que permite enviar e-mails, utilizando o protocolo SMTP
import ssl #Lib de protocolo para estabelecer uma conexão segura emtre cliente e servidor
from email.message import EmailMessage #Usada para criar e manipular mensagens de e-mail.
from Password import passwor #Importando senha do e-mail enviador, de outro arquivo Python.
import openpyxl #Lib utilizada para Abrir/Ler planilha do Excel.

workbook=openpyxl.load_workbook("usuarios.xlsx") #Carrega o arquivo Excel chamado "usuarios.xlsx"

worksheet = workbook.active #Obtém a planilha ativa da pasta de trabalho carregada e a armazena na variável worksheet. 

i = 2 # Indice utilizador para identificar a linha, que será inicializada.

while True:

    userexl = worksheet.cell(row=i, column=1).value #Pega e armaneza informação da Coluna 1, da linha conforme o Indice.
    email = worksheet.cell(row=i, column=2).value #Pega e armaneza da Coluna 2, da linha conforme o Indice .
    password = worksheet.cell(row=i, column=3).value #Pega e armaneza da Coluna 3, da linha conforme o Indice.

    user = str(userexl) #Transforma a informação da planilha em String.
    email_user = str(email) #Transforma a informação da planilha em String.
    password_user = str(password) #Transforma a informação da planilha em String.

    if userexl is None: #Verifica se a celula da planilha está vazia, caso esteja o laço é quebrado
        break

    elif userexl is not None: # Caso tenha alguma informação, o laço continua

        email_from = "" #E-mail de quem está mandando.
        pass_mail = passwor

                                                    #CORPO DO E-MAIL
        msg = EmailMessage()
        msg['Subject'] = ":: CREDENCIAIS DE ACESSO ::" #Assunto do e-mail
        msg['From'] = email_from #E-mail de quem está enviado
        msg['CC'] = email_from # Aqui você pode colocar outras pessoas em Copia no e-mail.
        msg['To'] = email #E-mail de quem receberá o e-mail.
        msg.set_content(f"Ola, {user}! \n \
                        \n Para melhor SEGURANÇA e RASTREABILIDADE de atividades em nossos sistemas, criamos logins para cada um de nossos colaboradores. \n \
                        Segue abaixo suas credenciais de acesso: \n \
                        \n Login:  {email} \nSenha: {password}")  # Mensagem que vai no corpo do e-mail.
        
        try: # Esse Try tentará executar esse bloco, que é responsável por enviar o e-mail.

            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp: #Protocolo SMTP com SSL para enviar o email. Aqui você coloca o smtp do seu e-mail e a porta ssl dele.
                    
                    print("Logando em seu email...")
                    smtp.login(email_from, pass_mail) #Loga no e-mail com usuario e senha.
                    print("Enviando email....")
                    smtp.send_message(msg) # Envia o e-mail.
                    print(f"Email enviado com sucesso, para {user}")

        except: # Caso ocorra algum erro com o bloco acima, ele exibira esse erro.

            print(" ###Algo deu errado em suas credenciais...###")
            break

        i += 1 #Acrescenta 1 no indice, para mudar de linha na planilha.
